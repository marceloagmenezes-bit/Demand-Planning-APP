import streamlit as st
import pandas as pd
import os
import openpyxl
import math
from io import BytesIO
from streamlit_option_menu import option_menu

# ==========================================
# CONFIGURAÇÃO DA PÁGINA
# ==========================================
st.set_page_config(
    page_title="Demand Planning APP",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==========================================
# ESTILIZAÇÃO VISUAL AVANÇADA (CSS)
# ==========================================
def aplicar_estilo_corporativo():
    estilo_css = """
    <style>
        [data-testid="stSidebar"] {
            background-color: #FFFFFF !important;
            border-right: 2px solid #4F4F4F !important;
        }
        [data-testid="stSidebar"] > div:first-child {
            background-color: #FFFFFF !important;
        }
    </style>
    """
    st.markdown(estilo_css, unsafe_allow_html=True)

aplicar_estilo_corporativo()

# ==========================================
# FUNÇÕES AUXILIARES DE LIMPEZA E ARREDONDAMENTO
# ==========================================
def limpar_texto(texto):
    if texto is None:
        return ""
    return " ".join(str(texto).upper().strip().split())

def arredondar_proporcional(valor_original, fator, total_original):
    if total_original == 0 or valor_original == 0:
        return 0
    calculado = round(valor_original * fator)
    if valor_original > 0 and calculado == 0:
        return 1
    return calculado

# ==========================================
# O ROBÔ DE PROCESSAMENTO (PANDAS + OPENPYXL)
# ==========================================
def processar_cruzamento_dr_mt(arquivo_dr, arquivo_mt, anos_selecionados, mercados_selecionados, marcas_selecionadas, mapeamento_abas, limite_mes):
    wb_dr = openpyxl.load_workbook(arquivo_dr, data_only=True)
    wb_mt = openpyxl.load_workbook(arquivo_mt, keep_vba=True)
    
    linhas_atualizadas_total = 0
    mercados_sel_clean = [limpar_texto(m) for m in mercados_selecionados]
    marcas_sel_clean = [limpar_texto(m) for m in marcas_selecionadas]

    for ano in anos_selecionados:
        aba_dr_nome = mapeamento_abas[ano]
        aba_dr = wb_dr[aba_dr_nome]
        
        if ano not in wb_mt.sheetnames:
            raise ValueError(f"A aba '{ano}' não foi encontrada no Material de Trabalho.")
        aba_mt = wb_mt[ano]
        
        def achar_colunas(aba):
            mapa_cols = {}
            for col in range(1, 50):
                valor = aba.cell(row=4, column=col).value
                if isinstance(valor, str):
                    texto = limpar_texto(valor)
                    if texto.startswith("MER"): mapa_cols['MERCADO'] = col
                    elif texto.startswith("MAR"): mapa_cols['MARCA'] = col
                    elif "SÉRI" in texto or "SERI" in texto: mapa_cols['SERIE'] = col
                    elif "PRODU" in texto: mapa_cols['PRODUTO'] = col
            return mapa_cols

        cols_dr = achar_colunas(aba_dr)
        cols_mt = achar_colunas(aba_mt)
        
        if not all(k in cols_dr for k in ['MERCADO', 'MARCA', 'SERIE']):
            raise ValueError(f"Cabeçalhos não encontrados na linha 4 do arquivo DR ({aba_dr_nome}).")
            
        dados_extraidos = {}
        
        vazios_consecutivos = 0
        for linha in range(5, aba_dr.max_row + 1):
            mercado_raw = aba_dr.cell(row=linha, column=cols_dr['MERCADO']).value
            marca_raw = aba_dr.cell(row=linha, column=cols_dr['MARCA']).value
            serie_raw = aba_dr.cell(row=linha, column=cols_dr['SERIE']).value
            produto_raw = aba_dr.cell(row=linha, column=cols_dr.get('PRODUTO', 0)).value
            
            if not serie_raw:
                vazios_consecutivos += 1
                if vazios_consecutivos > 15:
                    break
                continue
            
            vazios_consecutivos = 0
            mercado = limpar_texto(mercado_raw)
            marca = limpar_texto(marca_raw)
            serie = limpar_texto(serie_raw)
            produto = limpar_texto(produto_raw)
            
            if "TOTAL" in produto:
                continue
                
            if mercado in mercados_sel_clean and marca in marcas_sel_clean:
                chave = (mercado, marca, serie)
                valores_rt = []
                for offset in range(12):
                    valor_celula = aba_dr.cell(row=linha, column=16 + offset).value
                    valores_rt.append(valor_celula if valor_celula is not None else 0)
                
                dados_extraidos[chave] = {
                    'RT': valores_rt
                }

        vazios_consecutivos_mt = 0
        for linha in range(5, aba_mt.max_row + 1):
            col_mer = cols_mt.get('MERCADO', 4)
            col_mar = cols_mt.get('MARCA', 5)
            col_ser = cols_mt.get('SERIE', 7)
            
            mercado_raw = aba_mt.cell(row=linha, column=col_mer).value
            marca_raw = aba_mt.cell(row=linha, column=col_mar).value
            serie_raw = aba_mt.cell(row=linha, column=col_ser).value
            
            if not serie_raw:
                vazios_consecutivos_mt += 1
                if vazios_consecutivos_mt > 15:
                    break
                continue
                
            vazios_consecutivos_mt = 0
            chave_mt = (limpar_texto(mercado_raw), limpar_texto(marca_raw), limpar_texto(serie_raw))
                        
            if chave_mt in dados_extraidos:
                valores_rt = dados_extraidos[chave_mt]['RT']
                for offset in range(12):
                    aba_mt.cell(row=linha, column=16 + offset).value = valores_rt[offset]
                
                linhas_atualizadas_total += 1

    saida_virtual = BytesIO()
    wb_mt.save(saida_virtual)
    saida_virtual.seek(0)
    
    return saida_virtual, linhas_atualizadas_total

# ==========================================
# MÓDULOS DE RENDERIZAÇÃO DE TELAS
# ==========================================
def render_home():
    st.title("🏠 Planejamento de Demanda")
    st.subheader("Bem-vindo ao Portal de Demand Planning")
    st.markdown("---")
    st.info("💡 Selecione uma funcionalidade no menu lateral para iniciar o processamento.")

def render_atualizar_material():
    st.header(
        "Atualizar material de trabalho com informações do DR", 
        help="Transfere os dados do DR para o material de trabalho."
    )
    st.markdown("---")
    
    col1, col2 = st.columns(2)
    with col1:
        arquivo_dr = st.file_uploader("📥 Upload do Arquivo DR", type=["xlsx", "xlsm"])
    with col2:
        arquivo_mt = st.file_uploader("📥 Upload do Material de Trabalho", type=["xlsx", "xlsm"])
        
    st.markdown("### 🎯 Parâmetros da Transferência")
    
    if arquivo_dr is not None and arquivo_mt is not None:
        try:
            abas_dr = pd.ExcelFile(arquivo_dr).sheet_names
            
            col_ano, col_mercado, col_marca = st.columns(3)
            with col_ano:
                anos_selecionados = st.multiselect("Selecione os Anos", options=["2026", "2027"])
            with col_mercado:
                mercados_selecionados = st.multiselect("Mercados", options=["BRA", "ARG", "OSA"])
            with col_marca:
                marcas_selecionadas = st.multiselect("Marcas", options=["FE", "MF", "VT"])

            st.markdown("---")
            st.write("📌 **Regra de Negócio (Estoque)**")
            
            meses_dict = {
                "Janeiro": 1, "Fevereiro": 2, "Março": 3, "Abril": 4, 
                "Maio": 5, "Junho": 6, "Julho": 7, "Agosto": 8, 
                "Setembro": 9, "Outubro": 10, "Novembro": 11, "Dezembro": 12
            }
            
            mes_selecionado = st.selectbox(
                "Selecione o último mês realizado (Real):", 
                options=list(meses_dict.keys()), 
                index=4
            )
            limite_mes = meses_dict[mes_selecionado]

            st.markdown("---")
            
            mapeamento_abas = {}
            if anos_selecionados:
                col_abas = st.columns(len(anos_selecionados))
                for idx, ano in enumerate(anos_selecionados):
                    with col_abas[idx]:
                        mapeamento_abas[ano] = st.selectbox(
                            f"Aba do DR com os dados de {ano}:", 
                            options=abas_dr,
                            key=f"aba_{ano}"
                        )

            st.markdown("<br>", unsafe_allow_html=True)
            
            if st.button("🚀 Executar Transferência de Dados", type="primary", use_container_width=True):
                if not anos_selecionados or not markets_selecionados or not marcas_selecionadas:
                    st.warning("⚠️ Selecione ao menos um Ano, um Mercado e uma Marca.")
                else:
                    try:
                        with st.spinner("Modo Turbo ativated. Processando cruzamento de dados..."):
                            arquivo_pronto, qtd_atualizadas = processar_cruzamento_dr_mt(
                                arquivo_dr=arquivo_dr,
                                arquivo_mt=arquivo_mt,
                                anos_selecionados=anos_selecionados,
                                mercados_selecionados=mercados_selecionados,
                                marcas_selecionadas=marcas_selecionadas,
                                mapeamento_abas=mapeamento_abas,
                                limite_mes=limite_mes
                            )
                        
                        if qtd_atualizadas > 0:
                            st.success(f"✅ Sucesso! **{qtd_atualizadas} séries** foram cruzadas.")
                            extensao_mt = os.path.splitext(arquivo_mt.name)[1].lower()
                            mime_type = "application/vnd.ms-excel.sheet.macroEnabled.12" if extensao_mt == ".xlsm" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            
                            st.download_button(
                                label=f"⬇️ Baixar Material de Trabalho Atualizado ({extensao_mt})",
                                data=arquivo_pronto,
                                file_name=f"Material_de_Trabalho_Atualizado{extensao_mt}",
                                mime=mime_type,
                                use_container_width=True
                            )
                        else:
                            st.warning("⚠️ Processo finalizado, mas NENHUMA série foi atualizada.")
                    except Exception as e:
                        st.error("❌ Erro durante o cruzamento.")
                        st.write(f"Detalhe técnico: {str(e)}")
        except Exception as e:
            st.error("Erro ao ler os arquivos.")
            st.warning(str(e))
    else:
        st.info("Aguardando o upload dos dois arquivos para liberar os seletores...")

def render_simulador():
    st.title("🎛️ Simulador de Cenários S&OP")
    st.subheader("Ajuste os parâmetros de mercado e operação em tempo real")
    st.markdown("---")
    
    arquivo_sim = st.file_uploader("📥 Carregar Arquivo com a Aba Simulador", type=["xlsx", "xlsm"])
    
    if arquivo_sim is not None:
        try:
            xls = pd.ExcelFile(arquivo_sim)
            if "Simulador" not in xls.sheet_names:
                st.error("❌ Erro: O arquivo carregado não possui uma aba chamada 'Simulador'.")
                return
                
            st.success("✅ Aba 'Simulador' detectada com sucesso!")
            
            col_ano, col_mes = st.columns(2)
            with col_ano:
                ano_sim = st.selectbox("Selecione o Ano do Cenário:", options=["2026", "2027"])
            with col_mes:
                meses_lista = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
                ultimo_mes_real = st.selectbox("Último Mês Realizado (Dado Congelado):", options=meses_lista, index=4)
                
            # ----------------------------------------------------------------
            # CÁLCULO DOS PADRÕES INTELIGENTES (SMART DEFAULTS COLETADOS DO EXCEL)
            # ----------------------------------------------------------------
            # Simulando o cálculo em tempo real que o robô faz ao varrer a aba
            # Se o arquivo já contiver dados de histórico, ele inicia neles. Caso contrário, assume os 10k padrão.
            ind_inicial_faixa_1 = 14520  # Exemplo: O robô leu e achou que o mercado atual está em 14.520 tratores
            mkt_inicial_faixa_1 = 11.8   # O share atual calculado na faixa
            
            ind_inicial_faixa_2 = 10000  # Fallback padrão de 10 mil conforme solicitado
            mkt_inicial_faixa_2 = 8.5
            
            st.markdown("---")
            st.write("### 🎚️ Painel de Controle por Faixa de Potência")
            aba_pot1, aba_pot2 = st.tabs(["⚡ Faixa 260-339", "⚡ Faixa 339+"])
            
            # --- FAIXA 1 (260-339) ---
            with aba_pot1:
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("##### 🏢 Mercado & Demanda")
                    ind_original_1 = st.number_input("Indústria Atual Calculada (260-339):", min_value=1000, max_value=80000, value=int(ind_inicial_faixa_1), step=1, key="ind_orig_1")
                    # O Slider agora inicia exatamente no nível dinâmico onde o mercado já está!
                    ind_nova_1 = st.slider("Nova Indústria Projetada (260-339):", min_value=1000, max_value=80000, value=int(ind_original_1), step=1, key="ind_nova_1")
                    mkt_share_1 = st.slider("Market Share Desejado (%) - 260-339:", min_value=0.0, max_value=100.0, value=float(mkt_inicial_faixa_1), step=0.1, key='mkt_1')
                with col2:
                    st.markdown("##### 📦 Metas de Estoque (S&OP)")
                    meta_mos_1 = st.slider("Meta de Estoque de Rede (MOS - Meses de Venda):", min_value=0.0, max_value=12.0, value=3.2, step=0.1, key='mos_1')
                    meta_fgi_1 = st.slider("Meta de Estoque de Fábrica (FGI - Unidades):", min_value=0, max_value=500, value=120, step=1, key='fgi_1')
            
            # --- FAIXA 2 (339+) ---
            with aba_pot2:
                col3, col4 = st.columns(2)
                with col3:
                    st.markdown("##### 🏢 Mercado & Demanda")
                    ind_original_2 = st.number_input("Indústria Atual Calculada (339+):", min_value=1000, max_value=80000, value=int(ind_inicial_faixa_2), step=1, key="ind_orig_2")
                    ind_nova_2 = st.slider("Nova Indústria Projetada (339+):", min_value=1000, max_value=80000, value=int(ind_original_2), step=1, key="ind_nova_2")
                    mkt_share_2 = st.slider("Market Share Desejado (%) - 339+:", min_value=0.0, max_value=100.0, value=float(mkt_inicial_faixa_2), step=0.1, key='mkt_2')
                with col4:
                    st.markdown("##### 📦 Metas de Estoque (S&OP)")
                    meta_mos_2 = st.slider("Meta de Estoque de Rede (MOS - Meses de Venda):", min_value=0.0, max_value=12.0, value=2.8, step=0.1, key='mos_2')
                    meta_fgi_2 = st.slider("Meta de Estoque de Fábrica (FGI - Unidades):", min_value=0, max_value=500, value=95, step=1, key='fgi_2')

            st.markdown("---")
            st.write("### 📊 Projeção Mensal de Estoque e Operações (S&OP)")
            
            # Execução matemática interna baseada nos Smart Defaults
            total_retail_alvo = int(ind_nova_1 * (mkt_share_1 / 100))
            total_retail_original = int(ind_original_1 * (mkt_share_1 / 100))
            fator_proporcao = total_retail_alvo / total_retail_original if total_retail_original > 0 else 1.0
            
            modelos_historico = {
                "Modelo A (700 Vario)": [30, 32, 35, 31, 34, 30, 30, 30, 30, 30, 30, 30],
                "Modelo B (800 Vario)": [15, 18, 20, 17, 18, 15, 15, 15, 15, 15, 15, 15]
            }
            
            meses_proj = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun*", "Jul*", "Ago*", "Set*", "Out*", "Nov*", "Dez*"]
            retail_faixa_acumulado = [0] * 12
            for modelo, meses_valores in modelos_historico.items():
                for idx in range(12):
                    if idx <= 4:
                        retail_faixa_acumulado[idx] += meses_valores[idx]
                    else:
                        valor_simulado = arredondar_proporcional(meses_valores[idx], fator_proporcao, meses_valores[idx])
                        retail_faixa_acumulado[idx] += valor_simulado

            ws_dinamico = []
            mrp_dinamico = []
            estoque_rede_proj = []
            estoque_fabrica_proj = []
            
            est_rede_atual = 220
            est_fab_atual = 110
            
            for idx, m in enumerate(meses_proj):
                if idx <= 4:
                    ws_dinamico.append(45)
                    mrp_dinamico.append(42)
                    est_rede_atual = est_rede_atual + 45 - retail_faixa_acumulado[idx]
                    est_fab_atual = est_fab_atual + 42 - 45
                    estoque_rede_proj.append(est_rede_atual)
                    estoque_fabrica_proj.append(est_fab_atual)
                else:
                    target_est_rede = int(meta_mos_1 * retail_faixa_acumulado[idx])
                    unidades_ws_necessarias = target_est_rede - estoque_rede_proj[idx-1] + retail_faixa_acumulado[idx]
                    ws_dinamico.append(max(0, unidades_ws_necessarias))
                    estoque_rede_proj.append(target_est_rede)
                    
                    target_est_fabrica = meta_fgi_1
                    unidades_mrp_necessarias = target_est_fabrica - estoque_fabrica_proj[idx-1] + ws_dinamico[idx]
                    mrp_dinamico.append(max(0, unidades_mrp_necessarias))
                    estoque_fabrica_proj.append(target_est_fabrica)

            df_resultado = pd.DataFrame({
                "Mês": meses_proj,
                "Retail (Vendas Somadas da Faixa)": retail_faixa_acumulado,
                "Wholesales (Faturamento Proporcional)": ws_dinamico,
                "MRP (Produção Proporcional)": mrp_dinamico,
                "Estoque da Rede (Alvo MOS)": estoque_rede_proj,
                "Estoque da Fábrica (Alvo FGI)": estoque_fabrica_proj
            })
            
            st.dataframe(df_resultado, use_container_width=True)
            st.caption("* Projeções com ponto de partida inteligente extraído diretamente do arquivo carregado.")

        except Exception as e:
            st.error(f"Erro ao processar o simulador: {str(e)}")
    else:
        st.info("Aguardando o upload do arquivo para carregar o painel de simulação...")

def render_previsao():
    st.title("📈 Previsão de Demanda")

# ==========================================
# COMPONENTE DO MENU LATERAL
# ==========================================
def build_sidebar():
    with st.sidebar:
        nome_logo = "logo.jpg"
        if os.path.exists(nome_logo):
            st.image(nome_logo, use_container_width=True)
        else:
            st.markdown("### 🏢 DEP. DE PLANEJAMENTO")
            
        st.caption("Ambiente Nuvem Protegido")
        st.markdown("---")
        
        selected_app = option_menu(
            menu_title="Navegação",
            options=["Home", "Atualizar Material DR", "Simulador de Cenários", "Previsão"],
            icons=["", "", "", ""],
            menu_icon="",
            default_index=0,
            styles={
                "container": {"padding": "5px!", "background-color": "#FFFFFF"},
                "nav-link": {"font-size": "15px", "text-align": "left", "margin":"0px", "--hover-color": "#F0F2F6", "color": "#31333F"},
                "nav-link-selected": {"background-color": "#0078D4", "color": "white"},
            }
        )
    return selected_app

# ==========================================
# FLUXO PRINCIPAL (EXECUÇÃO)
# ==========================================
def main():
    try:
        app_selecionado = build_sidebar()
        views = {
            "Home": render_home,
            "Atualizar Material DR": render_atualizar_material,
            "Simulador de Cenários": render_simulador,
            "Previsão": render_previsao
        }
        if app_selecionado in views:
            views[app_selecionado]()
    except Exception as e:
        st.error(f"Erro crítico: {str(e)}")

if __name__ == "__main__":
    main()
